"""Per-year Excel writer — strictly the Belastingdienst per-trip fields.

Per the official rittenregistratie requirements, each trip row contains exactly
these fields (and nothing else):

- Datum — date
- Beginstand — begin odometer
- Eindstand — end odometer
- Vertrekadres — departure address
- Aankomstadres — arrival address
- Route — the driven route, only when it is not the most usual route
- Privé/zakelijk — private or business
- Privé-omrijkilometers — private detour km when a trip mixes business and
  private kilometres

No other per-trip columns are added.
"""
from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import Trip, TripType

HEADERS: List[str] = [
    "Datum",
    "Beginstand",
    "Eindstand",
    "Vertrekadres",
    "Aankomstadres",
    "Route",
    "Privé/zakelijk",
    "Privé-omrijkilometers",
]


def _pb(t: TripType) -> str:
    return "privé" if t is TripType.PRIVATE else "zakelijk"


class ExcelWriter:
    def __init__(self, data_dir: Path, car_id: str):
        self.data_dir = data_dir
        self.car_id = car_id
        self.data_dir.mkdir(parents=True, exist_ok=True)

    def _path_for_year(self, year: int) -> Path:
        return self.data_dir / f"trips-{self.car_id}-{year}.xlsx"

    def _open(self, year: int) -> tuple[Workbook, Worksheet, Path]:
        path = self._path_for_year(year)
        if path.exists():
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ritten"
            ws.append(HEADERS)
        return wb, ws, path

    def append_trip(self, trip: Trip) -> Path:
        year = trip.date.year
        wb, ws, path = self._open(year)
        ws.append(
            [
                trip.date.strftime("%Y-%m-%d"),
                trip.start_odo,
                trip.end_odo,
                trip.start_address,
                trip.end_address,
                trip.route,  # only populated when the usual route was not taken
                _pb(trip.trip_type),
                trip.private_detour_km or "",
            ]
        )
        wb.save(path)
        return path
