"""Vehicle-telemetry hooks: place resolution, auto-logging, naming, dedupe, HTTP."""
import importlib
from datetime import datetime
from pathlib import Path

import pytest

from rittenregistratie.config import Settings
from rittenregistratie.engine import Engine
from rittenregistratie.models import VehicleTripReport
from rittenregistratie.routes import RouteBook, haversine_m

HOME = (52.3702, 5.2147)
OFFICE = (52.0781, 4.3168)


def test_haversine_and_nearest():
    assert abs(haversine_m(52.0, 4.0, 52.0, 4.0)) < 1e-6
    rb = RouteBook({
        "home": {"address": "Home St 1", "lat": HOME[0], "lon": HOME[1]},
        "office": {"address": "Office 1", "lat": OFFICE[0], "lon": OFFICE[1], "radius_m": 100, "private": False},
        "noloc": {"address": "Somewhere"},
        "gym": {"address": "Gym 1", "lat": HOME[0] + 0.002, "lon": HOME[1], "private": True},  # ~220 m north of home
    }, {})
    name, addr, d = rb.nearest(HOME[0], HOME[1])
    assert (name, addr) == ("home", "Home St 1") and d < 1
    assert rb.nearest(HOME[0] + 0.001, HOME[1])[0] == "home"           # ~110 m: home (closer than gym)
    assert rb.nearest(HOME[0] + 0.0025, HOME[1])[0] == "gym"           # closer to gym
    assert rb.nearest(OFFICE[0] + 0.002, OFFICE[1]) is None            # 220 m > office's 100 m radius
    assert rb.nearest(OFFICE[0] + 0.0005, OFFICE[1])[0] == "office"
    assert rb.is_private_place("gym") and not rb.is_private_place("home") and not rb.is_private_place("nope")


@pytest.fixture
def settings(tmp_path: Path) -> Settings:
    (tmp_path / "config").mkdir()
    (tmp_path / "config" / "locations.yaml").write_text(
        f"Home: {{address: 'A St, Almere', lat: {HOME[0]}, lon: {HOME[1]}}}\n"
        f"spg: {{address: 'Waldorpstraat 3, Den Haag', lat: {OFFICE[0]}, lon: {OFFICE[1]}}}\n"
        "beach: {address: 'Strand 1, Zandvoort', lat: 52.37, lon: 4.53, private: true}\n"
    )
    (tmp_path / "config" / "routes.yaml").write_text("{}\n")
    (tmp_path / "config" / "cars.yaml").write_text(
        "mercedes:\n  label: 'Mercedes'\n  seed_address: Home\n  seed_odometer: 20000\n"
        "  phones: ['31612345678']\n"
    )
    return Settings(data_dir=tmp_path / "data", config_dir=tmp_path / "config",
                    trajectory_provider="maps_link", private_cap_plugin="warn",
                    whatsapp_app_secret="", hook_secret="s3cret", event_plugins="")


def _rows(eng, car="mercedes", year=2026):
    from openpyxl import load_workbook
    res = eng.exporter.generate(car, year)
    return list(load_workbook(res.path).active.iter_rows(values_only=True))[1:]


def test_known_place_by_zone_and_by_coordinates(settings):
    eng = Engine(settings)
    car = eng.cars.get("mercedes")
    r = eng.handle_vehicle_trip(car, VehicleTripReport(
        end_odo=20087, ended_at=datetime(2026, 9, 3, 9, 27), start_odo=20000,
        latitude=OFFICE[0] + 0.0004, longitude=OFFICE[1], zone="not_home", source="homeassistant"))
    assert r.logged and "Home -> spg (87 km, business)" in r and r.notice is False
    r2 = eng.handle_vehicle_trip(car, VehicleTripReport(
        end_odo=20171, ended_at=datetime(2026, 9, 3, 17, 22), start_odo=20087,
        latitude=HOME[0], longitude=HOME[1], zone="home"))
    assert "spg -> home (84 km" in r2
    rows = _rows(eng)
    assert [(a, b) for _, a, b, *_ in rows] == [(20000, 20087), (20087, 20171)]
    assert rows[0][4] == "Waldorpstraat 3, Den Haag" and rows[1][4] == "A St, Almere"
    # private flag from locations.yaml
    r3 = eng.handle_vehicle_trip(car, VehicleTripReport(end_odo=20240, latitude=52.37, longitude=4.53))
    assert "private" in r3 and _rows(eng)[-1][6] == "privé"


def test_unknown_place_is_logged_with_geocoded_address_then_named(settings, monkeypatch):
    import rittenregistratie.engine as eng_mod
    monkeypatch.setattr(eng_mod, "reverse_geocode", lambda lat, lon, key="": "Europaplein 24, Amsterdam")
    eng = Engine(settings)
    car = eng.cars.get("mercedes")
    r = eng.handle_vehicle_trip(car, VehicleTripReport(
        end_odo=20040, latitude=52.3412, longitude=4.8890, source="homeassistant"))
    assert r.logged and r.notice and "name <place>" in r
    row = _rows(eng)[-1]
    assert row[4] == "Europaplein 24, Amsterdam"
    ledger = eng.exporter.trips_for_year("mercedes", 2026)[-1]
    assert "end place not known" in ledger["note"] and ledger["raw_message"].startswith("[homeassistant] end_odo=20040 at 52.34120,4.88900")
    # teach the name; it gains coordinates so the next visit resolves by proximity
    assert eng.handle_user_command("name rai", "31612345678").startswith("Saved 'rai': Europaplein 24")
    assert eng.routebook.nearest(52.3412, 4.8890)[0] == "rai"
    assert "Nothing to name" in eng.handle_user_command("name again", "31612345678")
    r2 = eng.handle_vehicle_trip(car, VehicleTripReport(end_odo=20080, latitude=52.3413, longitude=4.8891))
    assert "-> rai (40 km" in r2 and r2.notice is False


def test_no_coordinates_falls_back_to_zone_text(settings, monkeypatch):
    import rittenregistratie.engine as eng_mod
    monkeypatch.setattr(eng_mod, "reverse_geocode", lambda lat, lon, key="": "")
    eng = Engine(settings)
    car = eng.cars.get("mercedes")
    eng.handle_vehicle_trip(car, VehicleTripReport(end_odo=20010, latitude=51.0, longitude=5.0))
    assert "51.00000,5.00000" in _rows(eng)[-1][4]  # bare coordinates when geocoding is unavailable
    eng.handle_vehicle_trip(car, VehicleTripReport(end_odo=20020, zone="work"))
    assert _rows(eng)[-1][4] == "work"


def test_ignored_when_not_above_last_and_gap_is_noted(settings):
    eng = Engine(settings)
    car = eng.cars.get("mercedes")
    assert eng.handle_vehicle_trip(car, VehicleTripReport(end_odo=20000, zone="home")) is None
    eng.handle_text("20050 spg", "31612345678", now=datetime(2026, 9, 3, 9, 0))
    assert eng.handle_vehicle_trip(car, VehicleTripReport(end_odo=20050, zone="spg")) is None  # already typed
    eng.handle_vehicle_trip(car, VehicleTripReport(end_odo=20150, start_odo=20070, zone="home"))
    note = eng.exporter.trips_for_year("mercedes", 2026)[-1]["note"]
    assert "ignition on was 20070; 20 km before this trip are not covered" in note
    assert _rows(eng)[-1][1:3] == (20050, 20150)  # chain, not the ignition reading


def test_manual_message_after_auto_trip_is_deduplicated(settings):
    eng = Engine(settings)
    car = eng.cars.get("mercedes")
    eng.handle_vehicle_trip(car, VehicleTripReport(end_odo=20087, zone="spg"))
    r = eng.handle_text("20087 spg", "31612345678", now=datetime(2026, 9, 3, 9, 30))
    assert "Already logged" in r and r.logged is False
    assert len(_rows(eng)) == 1
    # a genuinely new reading still logs
    assert "spg -> home" in eng.handle_text("20171 home", "31612345678", now=datetime(2026, 9, 3, 17, 30))


# --- HTTP hook ------------------------------------------------------------------

@pytest.fixture
def hook_client(tmp_path, monkeypatch):
    from rittenregistratie import events as ev

    class _EP:
        def __init__(self, name, fn): self.name, self._fn = name, fn
        def load(self): return self._fn

    def register(bus):
        def on_hook(payload):
            b = payload["body"]
            if b.get("ignition") != "off":
                return None
            return VehicleTripReport(end_odo=int(b["odometer"]), latitude=b.get("lat"),
                                     longitude=b.get("lon"), zone=b.get("zone", ""),
                                     start_odo=b.get("start_odometer"))
        bus.on(ev.hook_event("fakeha"), on_hook)

    monkeypatch.setattr(ev, "entry_points", lambda group: [_EP("fakeha", register)])
    cfg = tmp_path / "config"; cfg.mkdir()
    (cfg / "locations.yaml").write_text(
        f"Home: {{address: 'A St', lat: {HOME[0]}, lon: {HOME[1]}}}\nspg: {{address: 'W 3', lat: {OFFICE[0]}, lon: {OFFICE[1]}}}\n")
    (cfg / "routes.yaml").write_text("{}\n")
    (cfg / "cars.yaml").write_text(
        "mercedes:\n  label: M\n  seed_address: Home\n  seed_odometer: 20000\n  phones: ['31600000000', '31600000001']\n"
        "  event_plugins: [fakeha]\n"
        "van:\n  label: V\n  seed_address: Home\n  seed_odometer: 1\n  phones: ['31600000002']\n  event_plugins: []\n")
    for k, v in {"RIT_CONFIG_DIR": str(cfg), "RIT_DATA_DIR": str(tmp_path / "data"), "RIT_WHATSAPP_APP_SECRET": "",
                 "RIT_WHATSAPP_TOKEN": "tok", "RIT_WHATSAPP_PHONE_NUMBER_ID": "pid", "RIT_HOOK_SECRET": "s3cret",
                 "RIT_EVENT_PLUGINS": "", "RIT_REPLY_MODE": "text"}.items():
        monkeypatch.setenv(k, v)
    import rittenregistratie.config as config
    config._settings = None
    import rittenregistratie.main as main
    importlib.reload(main)
    sent = []

    async def fake_text(token, pid, to, text, graph_url=""):
        sent.append((to, str(text))); return True

    monkeypatch.setattr(main.whatsapp, "send_message", fake_text)
    from fastapi.testclient import TestClient
    return TestClient(main.app), sent


def test_hook_auth_and_routing(hook_client):
    client, sent = hook_client
    body = {"ignition": "off", "odometer": 20087, "lat": OFFICE[0], "lon": OFFICE[1], "zone": "not_home"}
    assert client.post("/hooks/fakeha/mercedes", json=body).status_code == 403
    h = {"X-Hook-Secret": "wrong"}
    assert client.post("/hooks/fakeha/mercedes", json=body, headers=h).status_code == 403
    h = {"X-Hook-Secret": "s3cret"}
    assert client.post("/hooks/fakeha/nocar", json=body, headers=h).status_code == 404
    r = client.post("/hooks/fakeha/van", json=body, headers=h)
    assert r.status_code == 404 and "not enabled" in r.json()["error"]
    r = client.post("/hooks/otherplugin/mercedes", json=body, headers=h)
    assert r.status_code == 404
    assert sent == []


def test_hook_logs_trip_and_notifies_all_car_phones(hook_client):
    client, sent = hook_client
    h = {"X-Hook-Secret": "s3cret"}
    r = client.post("/hooks/fakeha/mercedes", json={"ignition": "on", "odometer": 20000}, headers=h)
    assert r.json() == {"status": "ignored", "reason": "plugin returned no trip"}
    body = {"ignition": "off", "odometer": 20087, "lat": OFFICE[0], "lon": OFFICE[1], "zone": "not_home"}
    r = client.post("/hooks/fakeha/mercedes", json=body, headers=h)
    assert r.status_code == 200 and r.json()["status"] == "logged" and r.json()["end_odo"] == 20087
    assert sorted(to for to, _ in sent) == ["31600000000", "31600000001"]
    assert all("Home -> spg (87 km, business)" in t for _, t in sent)
    # same reading again: ignored, nobody notified twice
    sent.clear()
    r = client.post("/hooks/fakeha/mercedes", json=body, headers=h)
    assert r.json()["status"] == "ignored" and sent == []
    # the typed follow-up is recognised as already logged
    msg = {"entry": [{"changes": [{"value": {"messages": [
        {"from": "31600000000", "id": "w1", "type": "text", "text": {"body": "20087 spg"}}]}}]}]}
    client.post("/webhook", json=msg)
    assert any("Already logged" in t for _, t in sent)
