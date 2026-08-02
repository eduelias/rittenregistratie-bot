from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from rittenregistratie.config import Settings
from rittenregistratie.engine import Engine
from rittenregistratie.raw_ledger import RawLedger
from rittenregistratie.rebuild import rebuild


@pytest.fixture
def settings(tmp_path: Path) -> Settings:
    cfg = tmp_path / "config"
    cfg.mkdir()
    (cfg / "locations.yaml").write_text(
        "Home: {address: 'A St'}\nOffice: {address: 'B St'}\n"
    )
    (cfg / "routes.yaml").write_text("{}\n")
    (cfg / "cars.yaml").write_text(
        "car:\n  label: C\n  seed_address: Home\n  seed_odometer: 1000\n"
        "  phones: ['31600000000']\n"
    )
    return Settings(
        data_dir=tmp_path / "data", config_dir=cfg,
        trajectory_provider="maps_link", private_cap_plugin="warn",
        whatsapp_app_secret="",
    )


def test_ledger_records_every_trip(settings):
    eng = Engine(settings)
    eng.handle_text("1040 Office", "31600000000", now=datetime(2026, 1, 5, 9, 0))
    eng.handle_text("1090 Beach private", "31600000000", now=datetime(2026, 1, 6, 9, 0))
    ledger = RawLedger(settings.raw_ledger_file("car"))
    rows = ledger.read_all()
    assert len(rows) == 2
    assert rows[0].start_odo == 1000 and rows[0].end_odo == 1040
    assert rows[0].destination_raw == "Office" and not rows[0].is_private
    assert rows[1].is_private and rows[1].end_odo == 1090
    assert rows[1].raw_message == "1090 Beach private"


def test_ledger_is_append_only(settings):
    eng = Engine(settings)
    eng.handle_text("1040 Office", "31600000000", now=datetime(2026, 1, 5, 9, 0))
    path = settings.raw_ledger_file("car")
    first = path.read_text()
    eng.handle_text("1090 Office", "31600000000", now=datetime(2026, 1, 6, 9, 0))
    second = path.read_text()
    # the original content is still a prefix of the file (nothing rewritten)
    assert second.startswith(first)
    assert second.count("\n") == 2


def test_rebuild_reproduces_spreadsheet(settings, tmp_path):
    eng = Engine(settings)
    eng.handle_text("1040 Office", "31600000000", now=datetime(2026, 1, 5, 9, 0))
    eng.handle_text("1090 Office", "31600000000", now=datetime(2026, 1, 6, 9, 0))
    out = tmp_path / "rebuilt"
    n = rebuild("car", settings.data_dir, out)
    assert n == 2
    ws = load_workbook(out / "trips-car-2026.xlsx").active
    rows = [[ws.cell(r, c).value for c in range(1, 9)]
            for r in range(2, ws.max_row + 1)]
    assert rows[0][1] == 1000 and rows[0][2] == 1040  # begin/end odo
    assert rows[1][1] == 1040 and rows[1][2] == 1090
    assert rows[0][6] == "zakelijk"
