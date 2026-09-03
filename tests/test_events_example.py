"""The reference plugin in examples/ speaks the contract end to end."""
import sys
from pathlib import Path

import pytest

from rittenregistratie.config import Settings
from rittenregistratie.events import EXPORT_PRE_GENERATE, EventBus
from rittenregistratie.export import ExportService, main as export_main
from rittenregistratie.raw_ledger import RawLedger, RawTrip

EXAMPLE = Path(__file__).resolve().parents[1] / "examples" / "rittenregistratie-events-example"


@pytest.fixture
def example_register():
    sys.path.insert(0, str(EXAMPLE))
    try:
        from rittenregistratie_events_example import register
        yield register
    finally:
        sys.path.remove(str(EXAMPLE))
        sys.modules.pop("rittenregistratie_events_example", None)


@pytest.fixture
def ledger_settings(tmp_path):
    settings = Settings(data_dir=tmp_path / "data", config_dir=tmp_path / "config",
                        whatsapp_app_secret="")
    ledger = RawLedger(settings.raw_ledger_file("car"))
    ledger.append(RawTrip(timestamp="2026-02-01T09:00:00", start_odo=10, end_odo=50,
                          start_address="  Home   St 1 ", end_address="Office", destination_raw="office",
                          is_private=False, raw_message="50 office"))
    return settings


def test_example_plugin_round_trip(example_register, ledger_settings):
    bus = EventBus()
    example_register(bus)
    assert len(bus.handlers(EXPORT_PRE_GENERATE)) == 1
    svc = ExportService(ledger_settings, bus)
    res = svc.generate("car", 2026)
    assert res.rows == 1 and res.handlers == 1
    from openpyxl import load_workbook
    rows = list(load_workbook(res.path).active.iter_rows(values_only=True))
    assert rows[1][3] == "Home St 1"  # whitespace tidied by the plugin
    # Ledger untouched.
    assert RawLedger(ledger_settings.raw_ledger_file("car")).read_all()[0].start_address == "  Home   St 1 "


def test_export_cli_dry_run(ledger_settings, tmp_path, capsys):
    out = tmp_path / "out"
    rc = export_main(["car", "--no-plugins", "--out", str(out),
                      "--data-dir", str(ledger_settings.data_dir),
                      "--config-dir", str(ledger_settings.config_dir), "--json"])
    assert rc == 0
    text = capsys.readouterr().out
    assert "pass-through" in text and "2026: 1 rows (ledger had 1)" in text
    assert '"end_odo": 50' in text
    assert (out / "trips-car-2026.xlsx").exists()
    assert export_main(["nobody", "--no-plugins", "--data-dir", str(ledger_settings.data_dir),
                        "--config-dir", str(ledger_settings.config_dir)]) == 1
