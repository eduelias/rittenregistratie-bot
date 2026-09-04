from datetime import datetime
from pathlib import Path

import pytest

from rittenregistratie.config import Settings
from rittenregistratie.engine import Engine, EngineError
from openpyxl import load_workbook


def _rows(eng, year=2026, car="default_car"):
    """Rows of the on-demand export (spreadsheets are no longer written per trip)."""
    res = eng.exporter.generate(car, year)
    return list(load_workbook(res.path).active.iter_rows(values_only=True))


def _has_trips(eng, car="default_car"):
    return bool(eng.exporter.years(car))


@pytest.fixture
def settings(tmp_path: Path) -> Settings:
    (tmp_path / "config").mkdir()
    (tmp_path / "config" / "locations.yaml").write_text(
        "Home: {address: 'A St, Utrecht'}\nOffice: {address: 'B St, Amsterdam'}\n"
    )
    (tmp_path / "config" / "routes.yaml").write_text(
        "Home->Office: {expected_km: 40, variants: [38, 40, 43]}\n"
    )
    (tmp_path / "config" / "cars.yaml").write_text(
        "default_car:\n"
        "  label: 'Test car'\n"
        "  seed_address: Home\n"
        "  seed_odometer: 145000\n"
        "  phones: ['31612345678']\n"
    )
    return Settings(
        data_dir=tmp_path / "data",
        config_dir=tmp_path / "config",
        seed_address="Home",
        seed_odometer=145000,
        trajectory_provider="maps_link",
        
        private_cap_plugin="warn",
        whatsapp_app_secret="",
    )


def test_first_trip_uses_seed(settings):
    eng = Engine(settings)
    reply = eng.handle_text("145040 Office", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert "40 km" in reply
    assert "business" in reply
    rows = _rows(eng)
    assert rows[0][0] == "Datum"
    assert rows[1][1] == 145000  # Beginstand == seed
    assert rows[1][2] == 145040  # Eindstand


def test_no_route_lookup_by_default(settings):
    """The ledger cannot keep a route, so the core does not go looking for one."""
    eng = Engine(settings)
    reply = eng.handle_text("145050 Office", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert "route" not in reply.lower()
    assert not reply.notice, "a plain trip must not break through reaction mode"


def test_deviation_reported_when_asked_for(settings):
    settings.route_on_deviation = True
    eng = Engine(settings)
    reply = eng.handle_text("145050 Office", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert "deviation" in reply.lower() or "route" in reply.lower()
    assert reply.notice


def test_monotonic_enforced(settings):
    eng = Engine(settings)
    eng.handle_text("145040 Office", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    with pytest.raises(EngineError):
        eng.handle_text("145030 Home", "31612345678", now=datetime(2026, 1, 5, 18, 0))


def test_private_cap_warning(settings):
    eng = Engine(settings)
    # 600 km private trip exceeds the 500 km cap.
    reply = eng.handle_text("145600 Beach private", "31612345678", now=datetime(2026, 6, 1, 9, 0))
    assert "exceeds" in reply.lower()


def test_unknown_number_rejected(settings):
    from rittenregistratie.engine import UnknownCarError
    eng = Engine(settings)
    with pytest.raises(UnknownCarError):
        eng.handle_text("145040 Office", "31600000000", now=datetime(2026, 1, 5, 9, 0))


def test_unknown_destination_prompts_for_address(settings):
    eng = Engine(settings)
    reply = eng.handle_text("145050 Gym", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert "don't have an address" in reply.lower()
    # No trip written yet
    assert not _has_trips(eng)


def test_unknown_destination_resolved_by_text(settings):
    eng = Engine(settings)
    eng.handle_text("145050 Gym", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    reply = eng.handle_text("Sportlaan 1, Almere", "31612345678", now=datetime(2026, 1, 5, 9, 5))
    assert "Sportlaan 1" in reply
    rows = _rows(eng)
    assert rows[-1][4] == "Sportlaan 1, Almere"  # Aankomstadres
    # location learned for next time
    reply2 = eng.handle_text("145090 Gym", "31612345678", now=datetime(2026, 1, 6, 9, 0))
    assert "don't have an address" not in reply2.lower()


def test_unknown_destination_resolved_by_location(settings):
    eng = Engine(settings)
    eng.handle_text("145050 Gym", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    reply = eng.handle_location(
        "31612345678", 52.37, 4.90, address="Damrak 1, Amsterdam",
        now=datetime(2026, 1, 5, 9, 5),
    )
    assert "Damrak 1" in reply


def test_private_unknown_destination_not_prompted(settings):
    eng = Engine(settings)
    reply = eng.handle_text("145030 beach private", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert "don't have an address" not in reply.lower()
    assert "private" in reply.lower()


# --- address hardening -------------------------------------------------------

def test_destination_names_are_case_insensitive(settings):
    eng = Engine(settings)
    reply = eng.handle_text("145040 office", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert "don't have an address" not in reply.lower()
    rows = _rows(eng)
    assert rows[-1][4] == "B St, Amsterdam"


def test_learned_address_goes_to_data_not_config(settings):
    eng = Engine(settings)
    eng.handle_text("145050 Gym", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    eng.handle_text("Sportlaan 1, Almere", "31612345678", now=datetime(2026, 1, 5, 9, 5))
    config_text = (settings.config_dir / "locations.yaml").read_text()
    assert "Sportlaan" not in config_text
    learned = settings.learned_locations_file.read_text()
    assert "gym:" in learned and "Sportlaan 1, Almere" in learned
    # Survives a restart (new Engine reads the learned file).
    eng2 = Engine(settings)
    reply = eng2.handle_text("145090 GYM", "31612345678", now=datetime(2026, 1, 6, 9, 0))
    assert "don't have an address" not in reply.lower()


def test_new_trip_while_pending_is_not_swallowed_as_address(settings):
    eng = Engine(settings)
    eng.handle_text("145050 shr and back home", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    reply = eng.handle_text("145140 Office", "31612345678", now=datetime(2026, 1, 5, 9, 5))
    assert "dropped the pending trip" in reply.lower()
    assert "140 km" in reply  # pending dropped, so the trip starts at the seed 145000
    rows = _rows(eng)
    assert len(rows) == 2  # header + the Office trip only
    assert rows[1][1] == 145000 and rows[1][2] == 145140
    assert rows[1][4] == "B St, Amsterdam"
    # No bogus location learned.
    assert not settings.learned_locations_file.exists() or \
        "shr and back home" not in settings.learned_locations_file.read_text()


def test_cancel_drops_pending_trip(settings):
    eng = Engine(settings)
    eng.handle_text("145050 Gym", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    reply = eng.handle_text("cancel", "31612345678", now=datetime(2026, 1, 5, 9, 1))
    assert "cancelled" in reply.lower()
    assert not _has_trips(eng)
    # Next message is parsed as a normal trip again.
    reply2 = eng.handle_text("145040 Office", "31612345678", now=datetime(2026, 1, 5, 9, 2))
    assert "40 km" in reply2


def test_known_location_name_as_address_reuses_its_address(settings):
    eng = Engine(settings)
    eng.handle_text("145050 hq", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    reply = eng.handle_text("office", "31612345678", now=datetime(2026, 1, 5, 9, 1))
    assert "B St, Amsterdam" in reply
    rows = _rows(eng)
    assert rows[-1][4] == "B St, Amsterdam"  # not the literal 'office'


def test_numeric_only_reply_rejected_as_address(settings):
    eng = Engine(settings)
    eng.handle_text("145050 Gym", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    # Lower than the last odometer, so not a new trip; and no letters.
    reply = eng.handle_text("1234 56", "31612345678", now=datetime(2026, 1, 5, 9, 1))
    assert "doesn't look like an address" in reply.lower()
    assert not _has_trips(eng)


def test_reply_flags(settings):
    eng = Engine(settings)
    prompt = eng.handle_text("145050 Gym", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert prompt.logged is False
    learned = eng.handle_text("Sportlaan 1, Almere", "31612345678", now=datetime(2026, 1, 5, 9, 5))
    assert learned.logged is True and learned.notice is True
    plain = eng.handle_text("145090 Gym", "31612345678", now=datetime(2026, 1, 6, 9, 0))
    assert plain.logged is True and plain.notice is False


# --- storage model -----------------------------------------------------------

def test_no_spreadsheet_written_per_trip(settings):
    eng = Engine(settings)
    eng.handle_text("145040 Office", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert not list(settings.data_dir.rglob("*.xlsx"))
    assert settings.raw_ledger_file("default_car").exists()
    res = eng.exporter.generate("default_car", 2026)
    assert res.path.parent == settings.exports_dir and res.rows == 1


def test_parse_export_command(settings):
    eng = Engine(settings)
    now = datetime(2026, 3, 1, 9, 0)
    assert eng.parse_export_command("145040 Office", "31612345678") is None
    req = eng.parse_export_command("excel", "31612345678", now=now)
    assert req.car.car_id == "default_car" and req.years == [2026]
    assert eng.parse_export_command("/Export 2025", "31612345678").years == [2025]
    assert eng.parse_export_command("xlsx 2024 2025", "31612345678").years == [2024, 2025]
    with pytest.raises(EngineError, match="No trips"):
        eng.parse_export_command("excel all", "31612345678")
    eng.handle_text("145040 Office", "31612345678", now=datetime(2025, 6, 1))
    eng.handle_text("145080 Home", "31612345678", now=datetime(2026, 1, 5))
    assert eng.parse_export_command("excel all", "31612345678").years == [2025, 2026]
    # Anyone may name their own car; other cars need admin.
    assert eng.parse_export_command("excel default_car", "31612345678").car.car_id == "default_car"
    (settings.config_dir / "cars.yaml").write_text(
        (settings.config_dir / "cars.yaml").read_text()
        + "van:\n  label: Van\n  seed_address: Home\n  seed_odometer: 1\n  phones: ['31600000009']\n"
    )
    eng.reload_cars()
    with pytest.raises(EngineError, match="Unknown car 'van'"):
        eng.parse_export_command("excel van", "31612345678")
    settings.admin_numbers = "31612345678"
    admin_eng = Engine(settings)
    assert admin_eng.parse_export_command("excel van 2025", "31612345678").car.car_id == "van"
    with pytest.raises(EngineError, match="Unknown car"):
        admin_eng.parse_export_command("excel boat", "31612345678")


# --- many cars per phone, many phones per car ---------------------------------

@pytest.fixture
def multi(settings):
    (settings.config_dir / "cars.yaml").write_text(
        "car:\n  label: 'Family car'\n  seed_address: Home\n  seed_odometer: 145000\n"
        "  phones: ['31612345678', '31699999999']\n"
        "van:\n  label: 'Van'\n  seed_address: Office\n  seed_odometer: 302000\n"
        "  phones: ['31612345678']\n"
    )
    return Engine(settings)


def test_two_users_one_car_share_the_chain(multi):
    r1 = multi.handle_text("145040 Office", "31612345678", now=datetime(2026, 1, 5, 9, 0),
                           car=multi.cars.get("car"))
    assert "40 km" in r1
    # the other person continues from the same odometer/state
    r2 = multi.handle_text("145080 Home", "31699999999", now=datetime(2026, 1, 5, 18, 0))
    assert "[Family car] Office -> Home (40 km" in r2
    assert multi.exporter.generate("car", 2026).rows == 2


def test_multi_car_phone_must_choose(multi):
    with pytest.raises(EngineError, match="several cars"):
        multi.handle_text("145040 Office", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert not _has_trips(multi, "car") and not _has_trips(multi, "van")


def test_car_command_sets_active_car(multi):
    listing = multi.handle_user_command("cars", "31612345678")
    assert "Family car [car]" in listing and "Van [van]" in listing and "(active)" not in listing
    assert multi.handle_user_command("car van", "31612345678") == "Active car: Van [van]."
    reply = multi.handle_text("302050 Home", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert reply.startswith("[Van] Office -> Home (50 km")
    assert "(active)" in multi.handle_user_command("cars", "31612345678")
    # by label, case-insensitive
    assert multi.handle_user_command("car family CAR", "31612345678") == "Active car: Family car [car]."
    with pytest.raises(EngineError, match="Unknown car"):
        multi.handle_user_command("car boat", "31612345678")
    # a single-car phone: 'cars' works, listing marks the only car active
    assert "(active)" in multi.handle_user_command("cars", "31699999999")
    assert multi.handle_user_command("145040 Office", "31699999999") is None


def test_car_prefix_selects_car_for_one_message(multi):
    multi.handle_user_command("car car", "31612345678")
    reply = multi.handle_text("van 302050 Home", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    assert reply.startswith("[Van]")
    reply = multi.handle_text("145040 Office", "31612345678", now=datetime(2026, 1, 5, 10, 0))
    assert reply.startswith("[Family car]")  # active car unchanged by the prefix
    # a pending address on the van is answered with the same prefix
    prompt = multi.handle_text("van 302090 Gym", "31612345678", now=datetime(2026, 1, 5, 11, 0))
    assert "don't have an address" in prompt.lower()
    done = multi.handle_text("van Sportlaan 1, Almere", "31612345678", now=datetime(2026, 1, 5, 11, 1))
    assert done.startswith("[Van]") and "Sportlaan 1" in done


def test_export_command_own_car_by_name_without_admin(multi):
    multi.handle_text("van 302050 Home", "31612345678", now=datetime(2026, 1, 5, 9, 0))
    req = multi.parse_export_command("excel van 2026", "31612345678")
    assert req.car.car_id == "van" and req.years == [2026]
    req = multi.parse_export_command("excel 'Family car'".replace("'", ""), "31612345678")
    assert req.car.car_id == "car"
    with pytest.raises(EngineError, match="Unknown car"):
        multi.parse_export_command("excel boat", "31612345678")
    with pytest.raises(EngineError, match="several cars"):
        multi.parse_export_command("excel", "31612345678")
    multi.handle_user_command("car van", "31612345678")
    assert multi.parse_export_command("excel", "31612345678").car.car_id == "van"


def test_event_plugins_are_per_car(settings, monkeypatch):
    from rittenregistratie import events as ev

    class _EP:
        def __init__(self, name, fn): self.name, self._fn = name, fn
        def load(self): return self._fn

    def reg_mark(bus):
        bus.on(ev.EXPORT_PRE_GENERATE,
               lambda p: {**p, "trips": [{**t, "end_address": "MARKED"} for t in p["trips"]]})

    monkeypatch.setattr(ev, "entry_points", lambda group: [_EP("mark", reg_mark)])
    (settings.config_dir / "cars.yaml").write_text(
        "a:\n  label: A\n  seed_address: Home\n  seed_odometer: 100\n  phones: ['31600000001']\n"
        "  event_plugins: [mark]\n"
        "b:\n  label: B\n  seed_address: Home\n  seed_odometer: 100\n  phones: ['31600000002']\n"
        "  event_plugins: []\n"
        "c:\n  label: C\n  seed_address: Home\n  seed_odometer: 100\n  phones: ['31600000003']\n"
    )
    settings.event_plugins = ""  # global default: none
    eng = Engine(settings)
    for phone in ("31600000001", "31600000002", "31600000003"):
        eng.handle_text("140 Office", phone, now=datetime(2026, 1, 5, 9, 0))
    ra, rb, rc = (eng.exporter.generate(c, 2026) for c in ("a", "b", "c"))
    assert (ra.handlers, rb.handlers, rc.handlers) == (1, 0, 0)
    assert eng.loaded_event_plugins == {"a": ["mark"], "b": [], "c": []}
    assert _rows(eng, car="a")[1][4] == "MARKED"
    assert _rows(eng, car="b")[1][4] == "B St, Amsterdam"
    # global default applies to cars without the key
    settings.event_plugins = "mark"
    eng2 = Engine(settings)
    assert eng2.exporter.generate("c", 2026).handlers == 1
    assert eng2.exporter.generate("b", 2026).handlers == 0
