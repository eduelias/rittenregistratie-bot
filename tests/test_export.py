"""On-demand export: ledger -> pre_generate -> validate -> xlsx -> post_generate."""
from pathlib import Path

import pytest
from openpyxl import load_workbook

from rittenregistratie.config import Settings
from rittenregistratie.events import EXPORT_POST_GENERATE, EXPORT_PRE_GENERATE, EventBus
from rittenregistratie.export import ExportError, ExportService, validate_trips
from rittenregistratie.raw_ledger import RawLedger, RawTrip


def _trip(ts, start, end, dest="Office", private=False):
    return RawTrip(timestamp=ts, start_odo=start, end_odo=end, start_address="Home St 1",
                   end_address=dest, destination_raw=dest.lower(), is_private=private,
                   raw_message=f"{end} {dest}")


@pytest.fixture
def svc(tmp_path: Path):
    settings = Settings(data_dir=tmp_path / "data", config_dir=tmp_path / "config",
                        whatsapp_app_secret="")
    ledger = RawLedger(settings.raw_ledger_file("car"))
    ledger.append(_trip("2025-12-30T09:00:00", 100, 140))
    ledger.append(_trip("2026-01-05T09:00:00", 140, 180))
    ledger.append(_trip("2026-01-05T18:00:00", 180, 220, dest="Home St 1", private=True))
    bus = EventBus()
    return ExportService(settings, bus), bus, settings


def _rows(path):
    return list(load_workbook(path).active.iter_rows(values_only=True))


def test_years_and_filtering(svc):
    s, _, _ = svc
    assert s.years("car") == [2025, 2026]
    assert [t["end_odo"] for t in s.trips_for_year("car", 2026)] == [180, 220]
    assert s.years("nope") == []


def test_generate_identity_without_plugins(svc):
    s, _, settings = svc
    res = s.generate("car", 2026)
    assert res.path == settings.exports_dir / "trips-car-2026.xlsx"
    assert res.rows == 2 and res.handlers == 0
    rows = _rows(res.path)
    assert rows[0][0] == "Datum"
    assert rows[1][1:3] == (140, 180) and rows[1][6] == "zakelijk"
    assert rows[2][1:3] == (180, 220) and rows[2][6] == "privé"
    # No xlsx anywhere else, ledger untouched.
    assert not list(settings.data_dir.glob("*.xlsx"))
    assert len(RawLedger(settings.raw_ledger_file("car")).read_all()) == 3


def test_generate_replaces_previous_export(svc):
    s, _, _ = svc
    s.generate("car", 2026)
    res = s.generate("car", 2026)
    assert len(_rows(res.path)) == 3  # header + 2, not appended twice


def test_pre_generate_handler_output_is_what_gets_written(svc):
    s, bus, settings = svc
    seen = {}

    def handler(payload):
        seen.update(payload)
        trips = [dict(t) for t in payload["trips"]]
        trips[1]["is_private"] = False          # reclassify
        trips.append({                           # add a row
            "timestamp": "2026-01-06T09:00:00", "start_odo": 220, "end_odo": 260,
            "start_address": "Home St 1", "end_address": "Office", "is_private": False,
        })
        return {**payload, "trips": trips}

    posts = []
    bus.on(EXPORT_PRE_GENERATE, handler)
    bus.on(EXPORT_POST_GENERATE, posts.append)
    res = s.generate("car", 2026)
    assert seen["car_id"] == "car" and seen["year"] == 2026
    assert seen["data_dir"] == str(settings.data_dir)
    rows = _rows(res.path)
    assert res.rows == 3 and res.handlers == 1
    assert rows[2][6] == "zakelijk" and rows[3][1:3] == (220, 260)
    assert posts == [{"car_id": "car", "year": 2026, "path": str(res.path), "rows": 3}]
    # The ledger is not an output of the pipeline.
    assert [t.is_private for t in RawLedger(settings.raw_ledger_file("car")).read_all()] == [False, False, True]


def test_no_trips_is_an_error(svc):
    s, _, _ = svc
    with pytest.raises(ExportError, match="No trips"):
        s.generate("car", 2019)


@pytest.mark.parametrize("bad, msg", [
    ("not a list", "not a list"),
    ([1], "not an object"),
    ([{"timestamp": "2026-01-01T00:00:00"}], "missing field"),
    ([{"timestamp": "2026-01-01T00:00:00", "start_odo": "x", "end_odo": 1,
      "start_address": "a", "end_address": "b", "is_private": False}], "integers"),
    ([{"timestamp": "2026-01-01T00:00:00", "start_odo": 10, "end_odo": 5,
      "start_address": "a", "end_address": "b", "is_private": False}], "lower than start"),
    ([{"timestamp": "yesterday", "start_odo": 1, "end_odo": 5,
      "start_address": "a", "end_address": "b", "is_private": False}], "ISO"),
    ([{"timestamp": "2026-01-01T00:00:00", "start_odo": 1, "end_odo": 5,
      "start_address": "a", "end_address": "b", "is_private": "yes"}], "is_private"),
    ([{"timestamp": "2026-01-01T00:00:00", "start_odo": 1, "end_odo": 5,
      "start_address": "", "end_address": "b", "is_private": False}], "addresses"),
    ([{"timestamp": "2026-01-01T00:00:00", "start_odo": 1, "end_odo": 50,
      "start_address": "a", "end_address": "b", "is_private": False},
      {"timestamp": "2026-01-02T00:00:00", "start_odo": 40, "end_odo": 60,
      "start_address": "a", "end_address": "b", "is_private": False}], "chain broken"),
])
def test_validate_trips_rejects_bad_rows(bad, msg):
    with pytest.raises(ExportError, match=msg):
        validate_trips(bad)


def test_validate_sorts_and_allows_gaps():
    rows = validate_trips([
        {"timestamp": "2026-01-02T00:00:00", "start_odo": 60, "end_odo": 70,
         "start_address": "a", "end_address": "b", "is_private": False},
        {"timestamp": "2026-01-01T00:00:00", "start_odo": 1, "end_odo": 50,
         "start_address": "a", "end_address": "b", "is_private": True},
    ])
    assert [r["start_odo"] for r in rows] == [1, 60]


def test_bad_handler_payload_is_rejected(svc):
    s, bus, _ = svc
    bus.on(EXPORT_PRE_GENERATE, lambda p: ["oops"])
    with pytest.raises(ExportError, match="invalid payload"):
        s.generate("car", 2026)
    bus2 = EventBus()
    bus2.on(EXPORT_PRE_GENERATE, lambda p: {**p, "trips": []})
    with pytest.raises(ExportError, match="no trips"):
        ExportService(s.settings, bus2).generate("car", 2026)
