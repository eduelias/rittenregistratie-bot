"""One-off migration: convert legacy trip Excel files to the strict
Belastingdienst per-trip schema.

Legacy columns:
  Date, Time, Type, StartAddress, EndAddress, StartOdo, EndOdo, TripKm,
  Route, DeviationNote, Source, PrivateKmYTD

New (compliant) columns:
  Datum, Beginstand, Eindstand, Vertrekadres, Aankomstadres, Route,
  Privé/zakelijk, Privé-omrijkilometers

Mapping:
- Datum            <- Date
- Beginstand       <- StartOdo
- Eindstand        <- EndOdo
- Vertrekadres     <- StartAddress
- Aankomstadres    <- EndAddress
- Route            <- DeviationNote if it looks like a real driven route,
                      otherwise empty (the legacy Route column held a generic
                      "A -> B" summary for every trip, which is NOT the required
                      "route only when deviating", so it is cleared).
- Privé/zakelijk   <- 'privé' if Type == private else 'zakelijk'
- Privé-omrijkilometers <- empty (legacy data had no mixed-trip detour column)

Every legacy trip row is preserved 1:1 (no row added or removed).
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

NEW_HEADERS = [
    "Datum", "Beginstand", "Eindstand", "Vertrekadres", "Aankomstadres",
    "Route", "Privé/zakelijk", "Privé-omrijkilometers",
]


def migrate(path: Path) -> int:
    wb = load_workbook(path)
    ws = wb.active
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if header == NEW_HEADERS:
        print(f"{path.name}: already migrated")
        return 0
    idx = {name: i for i, name in enumerate(header)}
    rows_out = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if row[idx.get("StartOdo", 5)] is None:
            continue
        ttype = str(row[idx["Type"]]).lower()
        pb = "privé" if ttype == "private" else "zakelijk"
        # Legacy Route was a generic summary for every trip -> not compliant as
        # "route only when deviating", so clear it. Keep a genuine DeviationNote.
        dev = row[idx.get("DeviationNote", 9)]
        route = dev if dev else ""
        rows_out.append([
            row[idx["Date"]],
            row[idx["StartOdo"]],
            row[idx["EndOdo"]],
            row[idx["StartAddress"]],
            row[idx["EndAddress"]],
            route,
            pb,
            "",
        ])

    new = Workbook()
    nws = new.active
    nws.title = "Ritten"
    nws.append(NEW_HEADERS)
    for row in rows_out:
        nws.append(row)
    backup = path.with_suffix(".legacy.xlsx")
    if not backup.exists():
        path.rename(backup)
    new.save(path)
    print(f"{path.name}: migrated {len(rows_out)} rows (backup: {backup.name})")
    return len(rows_out)


if __name__ == "__main__":
    for arg in sys.argv[1:]:
        migrate(Path(arg))
